"""Render a report dict (from services.py) to Excel or PDF bytes."""
import io

from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .services import report_subtitle

_HEADER_FILL = PatternFill("solid", fgColor="B45309")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TITLE_FONT = Font(bold=True, size=14)
_OPENING_FILL = PatternFill("solid", fgColor="FEF3C7")  # light amber
_TOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")     # light gray
_BOLD = Font(bold=True)


def to_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = report["kind"].title()

    ncols = len(report["columns"])
    last_col = get_column_letter(ncols)

    # Title + subtitle.
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{report['shop'].name} — {report['title']}"
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = report_subtitle(report)
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    # Header row.
    header_row = 4
    for c, name in enumerate(report["columns"], start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    r = header_row + 1

    def write_row(cells, fill=None, bold=False):
        nonlocal r
        for c, value in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if fill:
                cell.fill = fill
            if bold:
                cell.font = _BOLD
        r += 1

    # Opening (amber) → entries → total (gray) → closing (gray). Every value
    # stays in its own column so debit/credit line up with the entry rows.
    if report.get("opening_row"):
        write_row(report["opening_row"], fill=_OPENING_FILL, bold=True)
    for row in report["rows"]:
        write_row(row)
    if report.get("total_row"):
        write_row(report["total_row"], fill=_OPENING_FILL, bold=True)
    if report.get("closing_row"):
        write_row(report["closing_row"], fill=_OPENING_FILL, bold=True)

    # Column widths.
    for c in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def to_pdf(report):
    html = render_to_string("reports/report.html", {
        "report": report,
        "subtitle": report_subtitle(report),
    })
    # Imported lazily so the app boots even where WeasyPrint's native libs
    # are unavailable (e.g. a minimal worker image without pango).
    from weasyprint import HTML

    return HTML(string=html).write_pdf()
